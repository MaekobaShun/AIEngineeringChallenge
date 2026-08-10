"""openpyxl + raw-XML based parser for .xlsx.

openpyxl gives us cell values and fill/font formatting directly, but two
things it doesn't expose at all are read generically from the underlying
OOXML parts:
  - AutoFilter criteria (which values a column is filtered to)
  - PivotTable definitions + the pivot cache's raw records

For PivotTables we don't try to replicate Excel's exact on-screen row
order; instead we rebuild the grouped aggregation ourselves (pandas
groupby over the pivot cache records, using the same row-fields/data-fields/
subtotal-function the table defines) so a question like "which group has
the highest average ALP" can be answered exactly. This generalizes to any
future pivot table, since row/data field selection is read from the file,
never hardcoded.

Large flat data sheets (e.g. a raw `train` sheet with thousands of rows) are
summarized (columns + row count) rather than dumped row-by-row -- the agent
should crunch those via the code-execution tool against the sibling CSV/
sheet directly, not via the text search index.
"""

from __future__ import annotations

import zipfile
from collections import Counter

import openpyxl
import pandas as pd
from lxml import etree
from openpyxl.utils import get_column_letter

from datastel_rag.ingest.models import Block, ParsedDocument

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_MAX_FULL_DUMP_ROWS = 200


def _fmt_num(v) -> str:
    try:
        return f"{float(v):.6g}"
    except (TypeError, ValueError):
        return str(v)


def _fill_color(cell) -> str | None:
    try:
        fill = cell.fill
        if fill is None or fill.fgColor is None:
            return None
        color = fill.fgColor
        if color.type == "rgb" and color.rgb and color.rgb != "00000000":
            return str(color.rgb)
        if color.type == "theme" and fill.patternType not in (None, "none"):
            return f"theme{color.theme}"
    except Exception:
        pass
    return None


def _sheet_blocks(ws, sheet_name: str) -> list[Block]:
    blocks = []
    nrows = ws.max_row or 0
    ncols = ws.max_column or 0
    if nrows == 0 or ncols == 0:
        return blocks

    if nrows > _MAX_FULL_DUMP_ROWS:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        blocks.append(
            Block(
                block_id=f"{sheet_name}/summary",
                kind="table_summary",
                text=f"シート'{sheet_name}': {nrows}行 x {ncols}列。列: {header}",
                location={"sheet": sheet_name},
                extra={"columns": header, "row_count": nrows},
            )
        )
        return blocks

    for row in ws.iter_rows(min_row=1, max_row=nrows, max_col=ncols):
        values = [c.value for c in row]
        if all(v is None for v in values):
            continue
        text = " | ".join("" if v is None else str(v) for v in values)
        colored = {}
        for c in row:
            fc = _fill_color(c)
            if fc:
                colored[f"{get_column_letter(c.column)}{c.row}"] = fc
        blocks.append(
            Block(
                block_id=f"{sheet_name}/row{row[0].row}",
                kind="table_row",
                text=text,
                location={"sheet": sheet_name, "row": row[0].row},
                extra={"fill_colors": colored} if colored else {},
            )
        )
    return blocks


def _parse_autofilters(zf: zipfile.ZipFile) -> dict[str, list[Block]]:
    """sheetN.xml -> autofilter description blocks (raw sheet-part names,
    matched to display names by the caller via workbook.xml/rels)."""
    out: dict[str, list[Block]] = {}
    for name in zf.namelist():
        if not (name.startswith("xl/worksheets/sheet") and name.endswith(".xml")):
            continue
        root = etree.fromstring(zf.read(name))
        af = root.find("m:autoFilter", _NS)
        if af is None:
            continue
        blocks = []
        for fc in af.findall("m:filterColumn", _NS):
            col_id = fc.get("colId")
            filters = fc.find("m:filters", _NS)
            values = [f.get("val") for f in filters.findall("m:filter", _NS)] if filters is not None else []
            custom = fc.find("m:customFilters", _NS)
            custom_desc = []
            if custom is not None:
                for cf in custom.findall("m:customFilter", _NS):
                    custom_desc.append(f"{cf.get('operator', '=')} {cf.get('val')}")
            desc = f"列インデックス{col_id}のフィルタ: 値={values}" if values else f"列インデックス{col_id}のフィルタ: 条件={custom_desc}"
            blocks.append(Block(block_id=f"autofilter/col{col_id}", kind="filter", text=desc, location={}, extra={"col_id": col_id, "values": values, "custom": custom_desc}))
        out[name] = blocks
    return out


def _cache_field_items(cache_field_el) -> list | None:
    shared = cache_field_el.find("m:sharedItems", _NS)
    if shared is None:
        return None
    items = []
    has_children = False
    for child in shared:
        has_children = True
        tag = etree.QName(child).localname
        if tag == "s":
            items.append(child.get("v"))
        elif tag == "n":
            items.append(float(child.get("v")))
        elif tag == "b":
            items.append(child.get("v") == "1")
        elif tag == "m":
            items.append(None)
        else:
            items.append(child.get("v"))
    return items if has_children else None


_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _resolve_cache_id_to_definition(zf: zipfile.ZipFile) -> dict[str, str]:
    """cacheId (as used by pivotTableN.xml) -> 'xl/pivotCache/pivotCacheDefinitionM.xml',
    resolved via workbook.xml's <pivotCaches> and workbook.xml.rels -- cacheId is a
    logical index, not necessarily the M in the definition's filename."""
    out: dict[str, str] = {}
    if "xl/workbook.xml" not in zf.namelist() or "xl/_rels/workbook.xml.rels" not in zf.namelist():
        return out
    wb_root = etree.fromstring(zf.read("xl/workbook.xml"))
    rels_root = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {rel.get("Id"): rel.get("Target") for rel in rels_root.findall(f"{{{_PKG_REL_NS}}}Relationship")}
    # search generically for pivotCache elements regardless of default-ns binding
    for pc in wb_root.iter():
        if etree.QName(pc).localname != "pivotCache":
            continue
        cache_id = pc.get("cacheId")
        rid = pc.get(f"{{{_R_NS}}}id")
        target = rid_to_target.get(rid)
        if cache_id is not None and target:
            out[cache_id] = "xl/" + target.lstrip("/")
    return out


def _parse_pivot_tables(zf: zipfile.ZipFile) -> list[Block]:
    blocks = []
    pt_names = [n for n in zf.namelist() if n.startswith("xl/pivotTables/pivotTable") and n.endswith(".xml")]
    cache_id_map = _resolve_cache_id_to_definition(zf)
    cache_def_names_fallback = sorted(n for n in zf.namelist() if n.startswith("xl/pivotCache/pivotCacheDefinition") and n.endswith(".xml"))

    for pi, pt_name in enumerate(pt_names):
        pt_root = etree.fromstring(zf.read(pt_name))
        cache_id = pt_root.get("cacheId", "0")

        cache_def_name = cache_id_map.get(cache_id)
        if cache_def_name is None or cache_def_name not in zf.namelist():
            try:
                cache_def_name = cache_def_names_fallback[int(cache_id)]
            except (IndexError, ValueError):
                cache_def_name = cache_def_names_fallback[0] if cache_def_names_fallback else None
        if cache_def_name is None:
            continue
        cache_idx_suffix = cache_def_name.rsplit("Definition", 1)[-1]  # e.g. "1.xml"
        records_name = f"xl/pivotCache/pivotCacheRecords{cache_idx_suffix}"

        cache_root = etree.fromstring(zf.read(cache_def_name))
        cache_fields = cache_root.findall(".//m:cacheField", _NS)
        field_names = [cf.get("name") for cf in cache_fields]
        field_items = [_cache_field_items(cf) for cf in cache_fields]

        if records_name not in zf.namelist():
            continue
        records_root = etree.fromstring(zf.read(records_name))
        records = []
        for r in records_root.findall("m:r", _NS):
            row = {}
            for (name, items), child in zip(zip(field_names, field_items), r):
                tag = etree.QName(child).localname
                if tag == "x":
                    idx = int(child.get("v"))
                    row[name] = items[idx] if items and 0 <= idx < len(items) else None
                elif tag == "n":
                    row[name] = float(child.get("v"))
                elif tag == "s":
                    row[name] = child.get("v")
                else:
                    row[name] = None
            records.append(row)
        if not records:
            continue
        df = pd.DataFrame(records)

        row_field_idxs = [int(f.get("x")) for f in pt_root.findall("m:rowFields/m:field", _NS) if int(f.get("x")) >= 0]
        row_field_names = [field_names[i] for i in row_field_idxs if i < len(field_names)]

        data_fields = pt_root.findall("m:dataFields/m:dataField", _NS)
        agg_map = {"average": "mean", "sum": "sum", "count": "count", "min": "min", "max": "max", "countNums": "count", "product": "prod", "stdDev": "std", "stdDevp": "std", "var": "var", "varp": "var"}

        if not row_field_names or not data_fields:
            continue

        agg_spec = {}
        display_names = {}
        for df_el in data_fields:
            fld_idx = int(df_el.get("fld"))
            col_name = field_names[fld_idx] if fld_idx < len(field_names) else f"field{fld_idx}"
            subtotal = df_el.get("subtotal", "sum")
            pandas_agg = agg_map.get(subtotal, "sum")
            agg_spec[col_name] = pandas_agg
            display_names[col_name] = df_el.get("name", col_name)

        try:
            grouped = df.groupby(row_field_names, dropna=False).agg(agg_spec).reset_index()
        except Exception as e:
            blocks.append(Block(block_id=f"pivot{pi}/error", kind="pivot_def", text=f"PivotTable再計算失敗: {e}", location={}))
            continue

        header = list(row_field_names) + [display_names[c] for c in agg_spec]
        lines = [f"PivotTable (グループ化キー: {row_field_names} / 集計: {display_names})", " | ".join(header)]
        for _, grow in grouped.iterrows():
            vals = [str(grow[c]) for c in row_field_names] + [_fmt_num(grow[c]) for c in agg_spec]
            lines.append(" | ".join(vals))

        blocks.append(
            Block(
                block_id=f"pivot{pi}",
                kind="pivot_table",
                text="\n".join(lines),
                location={"pivot_table_index": pi},
                extra={"row_fields": row_field_names, "data_fields": display_names, "num_groups": len(grouped)},
            )
        )
    return blocks


def parse_xlsx(abs_path: str, source_rel_path: str, project_key: str) -> ParsedDocument:
    doc = ParsedDocument(source_rel_path=source_rel_path, project_key=project_key, ext="xlsx")
    try:
        wb = openpyxl.load_workbook(abs_path, data_only=True)
    except Exception as e:
        doc.parse_errors.append(f"open failed: {e}")
        return doc

    for ws in wb.worksheets:
        doc.blocks.extend(_sheet_blocks(ws, ws.title))

    try:
        with zipfile.ZipFile(abs_path) as zf:
            af = _parse_autofilters(zf)
            for blocks in af.values():
                doc.blocks.extend(blocks)
            doc.blocks.extend(_parse_pivot_tables(zf))
    except Exception as e:
        doc.parse_errors.append(f"xml-level parse failed: {e}")

    doc.meta = {"sheet_names": wb.sheetnames}
    return doc
